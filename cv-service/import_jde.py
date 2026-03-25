#!/usr/bin/env python3
"""
JDE Revenue Import Script
=========================
Parses the JDE Excel revenue file and imports data into Supabase.

Usage:
    python import_jde.py                              # Full import
    python import_jde.py --dry-run                    # Parse only, no Supabase writes
    python import_jde.py --file /path/to/file.xlsx    # Custom file path

Expected Excel structure:
    - 13 sheets, each a revenue category
    - "ALL Category" sheet has all rows with a Category column
    - Row 1-2: headers/title
    - Row 3: column names (Customer Number, Customer Name, Jan-Dec dates, Total, Category)
    - Row 4+: data
"""

import json
import logging
import re
import ssl
import sys
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

EXCEL_FILE = "/Users/ai10/Downloads/Revenue Analysis 2025 By  Category.xlsx"
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
UNMATCHED_FILE = "/tmp/jde_unmatched.txt"

ALL_CATEGORY_SHEET = "ALL Category"

# Month columns in the Excel file (columns 3-14 typically)
# We detect these dynamically from row 3 headers

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("jde_import")

# ---------------------------------------------------------------------------
# Supabase REST helpers (using urllib — no external HTTP lib needed)
# ---------------------------------------------------------------------------

# Allow self-signed / unverified SSL for local dev if needed
_ssl_ctx = ssl.create_default_context()
_ssl_ctx.check_hostname = False
_ssl_ctx.verify_mode = ssl.CERT_NONE


def supabase_request(method: str, table: str, data: dict = None,
                     params: str = "", expect_rows: bool = False) -> Optional[list]:
    """Make a REST request to Supabase PostgREST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += f"?{params}"

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

    body = json.dumps(data).encode("utf-8") if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req, context=_ssl_ctx) as resp:
            raw = resp.read().decode("utf-8")
            if raw:
                return json.loads(raw)
            return []
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8") if e.fp else ""
        log.error(f"Supabase {method} {table} failed ({e.code}): {err_body}")
        return None
    except Exception as e:
        log.error(f"Supabase request error: {e}")
        return None


def supabase_get(table: str, params: str = "") -> list:
    """GET rows from a Supabase table."""
    result = supabase_request("GET", table, params=params)
    return result if result else []


def supabase_post(table: str, data: dict) -> Optional[list]:
    """INSERT a row into a Supabase table."""
    return supabase_request("POST", table, data=data)


def supabase_upsert(table: str, data: dict, on_conflict: str = "") -> Optional[list]:
    """UPSERT a row (insert or update on conflict)."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if on_conflict:
        url += f"?on_conflict={on_conflict}"

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=representation",
    }

    body = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")

    try:
        with urllib.request.urlopen(req, context=_ssl_ctx) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8") if e.fp else ""
        log.error(f"Supabase upsert {table} failed ({e.code}): {err_body}")
        return None
    except Exception as e:
        log.error(f"Supabase upsert error: {e}")
        return None


# ---------------------------------------------------------------------------
# Fuzzy matching
# ---------------------------------------------------------------------------

def normalize_name(name: str) -> str:
    """Normalize a name for matching: lowercase, strip special chars, collapse spaces."""
    if not name:
        return ""
    s = name.lower().strip()
    s = re.sub(r"[^a-z0-9\s]", "", s)   # remove non-alphanumeric except spaces
    s = re.sub(r"\s+", " ", s).strip()   # collapse whitespace
    return s


def normalize_aggressive(name: str) -> str:
    """Even more aggressive: remove ALL spaces too."""
    return normalize_name(name).replace(" ", "")


def fuzzy_match_tenant(jde_name: str, tenants: list) -> Optional[dict]:
    """
    Try to match a JDE customer name to a Supabase tenant.
    Returns the matched tenant dict or None.

    Strategy:
    1. Exact normalized match on brand_name or name
    2. Aggressive match (no spaces, no special chars)
    3. Contains match (JDE name contains tenant name or vice versa)
    4. First-word match
    """
    jde_norm = normalize_name(jde_name)
    jde_agg = normalize_aggressive(jde_name)

    if not jde_norm:
        return None

    # Pass 1: Exact normalized match
    for t in tenants:
        for field in ["brand_name", "name"]:
            val = t.get(field, "")
            if val and normalize_name(val) == jde_norm:
                return t

    # Pass 2: Aggressive match (no spaces)
    for t in tenants:
        for field in ["brand_name", "name"]:
            val = t.get(field, "")
            if val and normalize_aggressive(val) == jde_agg:
                return t

    # Pass 3: Contains match (either direction)
    for t in tenants:
        for field in ["brand_name", "name"]:
            val = t.get(field, "")
            if not val:
                continue
            t_norm = normalize_name(val)
            t_agg = normalize_aggressive(val)
            # Skip very short names to avoid false matches
            if len(t_agg) < 3 or len(jde_agg) < 3:
                continue
            if t_agg in jde_agg or jde_agg in t_agg:
                return t

    # Pass 4: First-word match (for multi-word names)
    jde_first = jde_norm.split()[0] if jde_norm.split() else ""
    if len(jde_first) >= 4:  # only if first word is meaningful
        for t in tenants:
            for field in ["brand_name", "name"]:
                val = t.get(field, "")
                if not val:
                    continue
                t_first = normalize_name(val).split()[0] if normalize_name(val).split() else ""
                if t_first and t_first == jde_first and len(t_first) >= 4:
                    return t

    return None


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def detect_month_columns(ws, header_row: int = 3) -> list:
    """
    Detect month columns from the header row.
    Returns list of (col_index, month, year) tuples.
    Month columns contain date values or month name strings.
    """
    months = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(header_row, col)
        val = cell.value
        if val is None:
            continue

        # If it's a datetime, extract month/year
        if isinstance(val, datetime):
            months.append((col, val.month, val.year))
            continue

        # If it's a string, try to parse month names
        if isinstance(val, str):
            val_lower = val.strip().lower()
            month_map = {
                "jan": 1, "feb": 2, "mar": 3, "apr": 4,
                "may": 5, "jun": 6, "jul": 7, "aug": 8,
                "sep": 9, "oct": 10, "nov": 11, "dec": 12,
                "january": 1, "february": 2, "march": 3, "april": 4,
                "june": 6, "july": 7, "august": 8, "september": 9,
                "october": 10, "november": 11, "december": 12,
            }
            for prefix, m in month_map.items():
                if val_lower.startswith(prefix):
                    # Try to extract year
                    year_match = re.search(r"20\d{2}", val)
                    year = int(year_match.group()) if year_match else 2025
                    months.append((col, m, year))
                    break

    return months


def detect_columns(ws, header_row: int = 3) -> dict:
    """
    Detect column positions from the header row.
    Returns dict with keys: customer_number, customer_name, total, category
    and their column indices.
    """
    cols = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(header_row, col).value
        if val is None:
            continue
        val_lower = str(val).strip().lower()

        if "customer number" in val_lower or "cust" in val_lower and "num" in val_lower:
            cols["customer_number"] = col
        elif "customer name" in val_lower or ("cust" in val_lower and "name" in val_lower):
            cols["customer_name"] = col
        elif val_lower == "total":
            cols["total"] = col
        elif val_lower == "category":
            cols["category"] = col

    return cols


def parse_all_category_sheet(wb) -> list:
    """
    Parse the ALL Category sheet.
    Returns list of dicts: {customer_number, customer_name, category, months: {(m,y): amount}}
    """
    if ALL_CATEGORY_SHEET not in wb.sheetnames:
        log.error(f"Sheet '{ALL_CATEGORY_SHEET}' not found! Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[ALL_CATEGORY_SHEET]
    log.info(f"Parsing '{ALL_CATEGORY_SHEET}': {ws.max_row} rows x {ws.max_column} cols")

    # Log first 3 rows for debugging
    for r in range(1, 4):
        row_vals = [repr(ws.cell(r, c).value) for c in range(1, min(ws.max_column + 1, 20))]
        log.info(f"  Row {r}: {row_vals}")

    # Detect structure
    cols = detect_columns(ws, header_row=3)
    month_cols = detect_month_columns(ws, header_row=3)

    log.info(f"  Detected columns: {cols}")
    log.info(f"  Detected {len(month_cols)} month columns: {[(m, y) for _, m, y in month_cols]}")

    if "customer_name" not in cols:
        # Try header row 2
        cols = detect_columns(ws, header_row=2)
        month_cols = detect_month_columns(ws, header_row=2)
        log.info(f"  Retried with row 2: columns={cols}, months={len(month_cols)}")

    if "customer_name" not in cols:
        log.error("Could not detect 'Customer Name' column!")
        # Fallback: assume standard layout
        cols = {"customer_number": 1, "customer_name": 2, "total": 15}
        month_cols = [(c, c - 2, 2025) for c in range(3, 15)]  # cols 3-14 = Jan-Dec
        log.warning(f"  Using fallback layout: {cols}")

    # Parse data rows (starting from row 4)
    rows = []
    data_start = 4
    cust_name_col = cols["customer_name"]
    cust_num_col = cols.get("customer_number", 1)
    cat_col = cols.get("category")

    for r in range(data_start, ws.max_row + 1):
        name = ws.cell(r, cust_name_col).value
        if not name or not str(name).strip():
            continue
        name = str(name).strip()

        # Skip summary/total rows
        if name.lower() in ("total", "grand total", "sum"):
            continue

        cust_num = ws.cell(r, cust_num_col).value
        category = ws.cell(r, cat_col).value if cat_col else None

        months_data = {}
        for col_idx, month, year in month_cols:
            val = ws.cell(r, col_idx).value
            if val is not None:
                try:
                    amount = float(val)
                    if amount != 0:
                        months_data[(month, year)] = amount
                except (ValueError, TypeError):
                    pass

        rows.append({
            "customer_number": str(cust_num).strip() if cust_num else None,
            "customer_name": name,
            "category": str(category).strip() if category else None,
            "months": months_data,
        })

    log.info(f"  Parsed {len(rows)} tenant rows from ALL Category")
    return rows


def parse_category_sheets(wb) -> dict:
    """
    Parse individual category sheets for revenue breakdown.
    Returns: {customer_name: {(month, year): {category: amount}}}
    """
    breakdown = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == ALL_CATEGORY_SHEET:
            continue

        ws = wb[sheet_name]
        category = sheet_name.strip()
        log.info(f"Parsing category sheet: '{category}' ({ws.max_row} rows)")

        cols = detect_columns(ws, header_row=3)
        month_cols = detect_month_columns(ws, header_row=3)

        if "customer_name" not in cols:
            cols = detect_columns(ws, header_row=2)
            month_cols = detect_month_columns(ws, header_row=2)

        if "customer_name" not in cols:
            # Fallback
            cols = {"customer_number": 1, "customer_name": 2}
            month_cols = [(c, c - 2, 2025) for c in range(3, 15)]

        cust_name_col = cols["customer_name"]
        data_start = 4

        for r in range(data_start, ws.max_row + 1):
            name = ws.cell(r, cust_name_col).value
            if not name or not str(name).strip():
                continue
            name = str(name).strip()
            if name.lower() in ("total", "grand total", "sum"):
                continue

            if name not in breakdown:
                breakdown[name] = {}

            for col_idx, month, year in month_cols:
                val = ws.cell(r, col_idx).value
                if val is not None:
                    try:
                        amount = float(val)
                        if amount != 0:
                            key = (month, year)
                            if key not in breakdown[name]:
                                breakdown[name][key] = {}
                            breakdown[name][key][category] = amount
                    except (ValueError, TypeError):
                        pass

    return breakdown


# ---------------------------------------------------------------------------
# Import logic
# ---------------------------------------------------------------------------

def fetch_tenants() -> list:
    """Fetch all tenants from Supabase."""
    tenants = supabase_get("tenants", "select=id,name,brand_name,category&limit=1000")
    log.info(f"Fetched {len(tenants)} tenants from Supabase")
    return tenants


def fetch_leases() -> dict:
    """Fetch active leases, return {tenant_id: lease_id} mapping."""
    leases = supabase_get("leases", "select=id,tenant_id&status=eq.active&limit=1000")
    mapping = {}
    for l in leases:
        tid = l.get("tenant_id")
        if tid:
            mapping[tid] = l["id"]
    log.info(f"Fetched {len(mapping)} active leases")
    return mapping


def import_revenue(rows: list, breakdown: dict, dry_run: bool = False):
    """Match JDE rows to tenants and import into Supabase."""
    tenants = fetch_tenants()
    lease_map = fetch_leases()

    matched = 0
    unmatched = 0
    transactions_created = 0
    unmatched_names = []
    match_log = []

    for row in rows:
        jde_name = row["customer_name"]
        tenant = fuzzy_match_tenant(jde_name, tenants)

        if not tenant:
            unmatched += 1
            unmatched_names.append(f"{row.get('customer_number', '?')} | {jde_name}")
            log.warning(f"  UNMATCHED: {jde_name}")
            continue

        matched += 1
        tenant_id = tenant["id"]
        tenant_brand = tenant.get("brand_name") or tenant.get("name")
        lease_id = lease_map.get(tenant_id)

        match_log.append(f"  MATCHED: '{jde_name}' -> '{tenant_brand}' (tenant_id={tenant_id})")
        log.info(match_log[-1])

        if not lease_id:
            log.warning(f"    No active lease for tenant {tenant_brand} ({tenant_id}), skipping transactions")
            continue

        # Import monthly transactions
        for (month, year), amount in row["months"].items():
            if dry_run:
                log.info(f"    [DRY RUN] Would create: {month}/{year} = EGP {amount:,.2f}")
                transactions_created += 1
                continue

            tx_data = {
                "lease_id": lease_id,
                "period_month": month,
                "period_year": year,
                "amount_due": amount,
                "amount_paid": amount,
                "status": "paid",
                "source": "jde_import",
            }

            result = supabase_upsert(
                "rent_transactions",
                tx_data,
                on_conflict="lease_id,period_month,period_year",
            )

            if result is not None:
                transactions_created += 1
            else:
                # If upsert fails (maybe no unique constraint), try plain insert
                result = supabase_post("rent_transactions", tx_data)
                if result is not None:
                    transactions_created += 1
                else:
                    log.error(f"    Failed to create tx: {month}/{year} EGP {amount}")

        # Import category breakdown if available
        jde_breakdown = breakdown.get(jde_name, {})
        if jde_breakdown and not dry_run:
            import_category_breakdown(tenant_id, lease_id, tenant_brand, jde_breakdown)
        elif jde_breakdown and dry_run:
            for (month, year), cats in jde_breakdown.items():
                cats_str = ", ".join(f"{c}: EGP {a:,.2f}" for c, a in cats.items())
                log.info(f"    [DRY RUN] Breakdown {month}/{year}: {cats_str}")

    # Summary
    print("\n" + "=" * 60)
    print("JDE REVENUE IMPORT SUMMARY")
    print("=" * 60)
    print(f"Total tenant rows parsed:    {len(rows)}")
    print(f"Tenants matched:             {matched}")
    print(f"Tenants unmatched:           {unmatched}")
    print(f"Transactions created:        {transactions_created}")
    if dry_run:
        print("MODE:                        DRY RUN (no writes)")
    print("=" * 60)

    # Save unmatched
    if unmatched_names:
        with open(UNMATCHED_FILE, "w") as f:
            f.write("JDE Unmatched Tenants\n")
            f.write(f"Generated: {datetime.now().isoformat()}\n")
            f.write(f"Total unmatched: {len(unmatched_names)}\n")
            f.write("-" * 60 + "\n")
            for name in sorted(unmatched_names):
                f.write(name + "\n")
        print(f"\nUnmatched tenants saved to: {UNMATCHED_FILE}")

    return matched, unmatched, transactions_created


def import_category_breakdown(tenant_id: str, lease_id: str, tenant_brand: str,
                               breakdown_data: dict):
    """
    Import per-category revenue breakdown for a tenant.
    Stores in tenant_sales_reported or a revenue_breakdown table.

    breakdown_data: {(month, year): {category_name: amount}}
    """
    for (month, year), categories in breakdown_data.items():
        # Build a JSON summary of category breakdown
        breakdown_json = json.dumps(categories)
        total = sum(categories.values())

        # Store as tenant_sales_reported with breakdown in notes
        report_data = {
            "lease_id": lease_id,
            "tenant_id": tenant_id,
            "period_month": month,
            "period_year": year,
            "reported_revenue_egp": total,
            "submission_date": datetime.now().strftime("%Y-%m-%d"),
            "verified": True,
            "verification_notes": f"JDE import. Breakdown: {breakdown_json}",
        }

        result = supabase_upsert(
            "tenant_sales_reported",
            report_data,
            on_conflict="tenant_id,period_month,period_year",
        )

        if result is None:
            # Fallback to plain insert
            supabase_post("tenant_sales_reported", report_data)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    import argparse

    parser = argparse.ArgumentParser(description="Import JDE revenue data into Supabase")
    parser.add_argument("--file", "-f", default=EXCEL_FILE, help="Path to Excel file")
    parser.add_argument("--dry-run", "-n", action="store_true", help="Parse only, no Supabase writes")
    parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    excel_path = Path(args.file)
    if not excel_path.exists():
        log.error(f"Excel file not found: {excel_path}")
        sys.exit(1)

    log.info(f"Loading Excel file: {excel_path}")
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    log.info(f"Sheets found: {wb.sheetnames}")

    # Step 1: Parse ALL Category sheet
    rows = parse_all_category_sheet(wb)

    # Step 2: Parse individual category sheets for breakdown
    log.info("\nParsing individual category sheets for breakdown...")
    breakdown = parse_category_sheets(wb)
    log.info(f"Category breakdown parsed for {len(breakdown)} tenants")

    # Step 3: Import into Supabase
    if args.dry_run:
        log.info("\n--- DRY RUN MODE ---")

    import_revenue(rows, breakdown, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
